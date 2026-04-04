#!/usr/bin/env python3

from ibapi.contract import Contract
from ibapi.common import *
import threading
import time
from datetime import datetime
import openpyxl
import pandas as pd
import requests
from bs4 import BeautifulSoup
import argparse
from openpyxl.styles import Font
from ib_connector_base import IBConnector
import utils


class IBAccountInfo(IBConnector):

    EXCEL_FILES_DIR = "ExcelFiles/"

    def __init__(self):
        super().__init__()
        self.account_balance_fields = ["NetLiquidationByCurrency",
                                       "StockMarketValue",
                                       "TotalCashBalance",
                                       "NetDividend",
                                       "UnrealizedPnL"]
        account_balance_dict = {field: {"USD": 0.0, "ILS": 0.0} for field in self.account_balance_fields}
        self.account_balance_info = pd.DataFrame(account_balance_dict)
        self.account_balance_file = 'account_info.xlsx'
        self.deposits_file = 'Deposits.xlsx'
        self.sub_accounts = []  # Store sub-account identifiers
        self.account_data_received = threading.Event()  # Event to signal data is received
        self.current_account = None  # Track the current account being processed
        self.account_desc = {"U13834548": "ETFS", "U17039071": "SGOV", "U19254019": "Stocks"}

    def managedAccounts(self, accountsList):
        """
        Callback to get the list of sub-accounts under the master account.
        """
        self.sub_accounts = accountsList.split(",")[:-1]
        print(f"Managed accounts: {self.sub_accounts}")
        self.account_data_received.set()  # Signal that the sub-account data has been received

    def contractDetails(self, reqId: int, contractDetails):
        print(f"ContractDetails. ReqId: {reqId}, Contract: {contractDetails}")

    def tickPrice(self, reqId: int, tickType, price, attrib):
        bid_tick = 9  # Assuming tickType 9 is the bid price
        if reqId == self.req_ids['exchange_rate'] and tickType == bid_tick:
            self.account_balance_info["IbExchangeRate"] = price
            self.exchange_rate_received.set()

    def accountSummary(self, reqId, account: str, tag: str, value: str, currency: str):
        """
        Handle account summary data for each account.
        """
        if account not in self.account_balance_info:
            # Initialize account balance info for the account if it doesn't exist yet
            self.account_balance_info[account] = {field: {"USD": 0.0, "ILS": 0.0} for field in
                                                  self.account_balance_fields}

        if tag in self.account_balance_fields:
            # Handle base currency (assuming it's USD)
            if currency == "BASE" or currency == "USD":
                self.account_balance_info[account][tag]["USD"] = round(float(value), 2)
            elif currency == "ILS":
                self.account_balance_info[account][tag]["ILS"] = round(float(value), 2)

        #print(f"Account: {account}, Tag: {tag}, Value: {value}, Currency: {currency}")

    def request_account_summary_from_api(self, account_id):
        """
        Fetch account summary for a specific sub-account.
        """
        self.current_account = account_id
        req_id = self.req_ids["account_summary"]
        self.reqAccountSummary(req_id, "All", "$LEDGER")
        time.sleep(2)
        self.cancelAccountSummary(req_id)    

    def request_ib_exchange_rate(self):
        """
        fetch IB USD to ILS exchange rate
        note that the retrieved rate is not in real-time thus cannot be depended on
        :return:
        """
        contract = Contract()
        contract.symbol = "USD"
        contract.secType = "CASH"
        contract.currency = "ILS"
        contract.exchange = "IDEALPRO"
        req_id = self.req_ids['exchange_rate']
        self.reqMktData(req_id, contract, "", False, False, [])
        self.exchange_rate_received.wait(timeout=10)
        self.cancelMktData(req_id)  # Cancel the request to stop receiving updates

    def get_total_deposits(self):
        # get total ILS deposits since inception from the manually updated file
        deposits_df = pd.read_excel(self.EXCEL_FILES_DIR + self.deposits_file)
        return deposits_df[deposits_df.Amount > 0]["Amount"].sum()  # filter out withdrawals (have negative values)

    def get_account_info(self, write_to_excel):
        """
        Retrieve account information for all sub-accounts.
        """
        # Wait for managedAccounts to populate sub_accounts
        self.account_data_received.wait(timeout=10)  # Wait until managedAccounts is called

        # Prepare a DataFrame for storing aggregated values (sum of all sub-accounts)
        sum_data = {field: {"USD": 0.0, "ILS": 0.0} for field in self.account_balance_fields}
        sum_df = pd.DataFrame(sum_data).T  # Transpose to get correct shape

        # Get the latest USD to ILS exchange rate
        exchange_rate = utils.get_usd_to_ils_exchange_rate()

        # For each sub-account, retrieve and process account information
        self.account_balance_info = {}  # Clear any previous data
        for account in self.sub_accounts:
            print(f"Fetching account information for: {account}")

            # Initialize the account data structure before fetching the account summary
            self.account_balance_info[account] = {field: {"USD": 0.0, "ILS": 0.0} for field in
                                                  self.account_balance_fields}

            # Fetch account summary for the current sub-account
            self.request_account_summary_from_api(account)

            # Update ILS values based on the latest exchange rate
            for tag in self.account_balance_fields:
                # Convert USD to ILS using the fetched exchange rate
                usd_value = self.account_balance_info[account][tag]["USD"]
                self.account_balance_info[account][tag]["ILS"] = round(usd_value * exchange_rate, 2)

            # Aggregate the values to the sum DataFrame
            for tag in self.account_balance_fields:
                sum_df.at[tag, "USD"] += self.account_balance_info[account][tag]["USD"]
                sum_df.at[tag, "ILS"] += self.account_balance_info[account][tag]["ILS"]

        # Optionally write to Excel
        if write_to_excel:
            self.write_account_info_to_excel(sum_df, self.account_balance_info, exchange_rate)

        return sum_df, self.account_balance_info

    def write_account_info_to_excel(self, sum_df, account_info, exchange_rate):
        """
        Write account info into an Excel file. Three rows will be written: sum of accounts and individual account details.
        """
        # Define the headers for the Excel file
        headers = [
            ('Date', 1),  # col 1
            ('Exchange Rate', 1),  # col 2
            ('Type', 1),  # col 3
            ('Net Liquidation', 2),  # cols 4, 5
            ('Stock Market Value', 2),  # cols 6, 7
            ('Total Cash Balance', 2),  # cols 8, 9
            ('Net Dividend', 2),  # cols 10, 11
            ('IB Unrealized USD PnL', 3),  # cols 12, 13, 14
            ('Total ILS Deposits', 1),  # col 15 (only filled in the sum row)
            ('Unrealized ILS PnL', 3),  # cols 16, 17, 18
            ('Unrealized USD PnL', 3)  # cols 19, 20, 21
        ]

        # Load the Excel workbook or create a new one if it doesn't exist
        try:
            workbook = openpyxl.load_workbook(self.EXCEL_FILES_DIR + self.account_balance_file)
            sheet = workbook.active

            # Check if there is already data in the file (beyond headers)
            if sheet.max_row > 2:
                # Explicitly append a blank row by writing an empty string to each column
                sheet.append(["" for _ in range(sheet.max_column)])

        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            self.write_headers(sheet, headers)

        # Define the styles in correct order
        usd_currency_style, nis_currency_style, percentage_style = self.define_excel_styles(workbook)

        # Get current date and time
        curr_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Write the sum of all accounts (first row)
        self.write_row_to_excel(sheet, curr_datetime, exchange_rate, "Sum of Accounts", sum_df, usd_currency_style,
                                nis_currency_style, percentage_style, total_ils_deposits=self.get_total_deposits())

        # Write individual account data (second and third rows)
        for account, df in account_info.items():
            account_name = account + " " + self.account_desc[account]
            self.write_row_to_excel(sheet, None, None, account_name, df, usd_currency_style,
                                    nis_currency_style, percentage_style)

        # Save the workbook
        workbook.save(self.EXCEL_FILES_DIR + self.account_balance_file)

    def write_row_to_excel(self, sheet, date, exchange_rate, row_type, df, usd_style, nis_style, percentage_style,
                           total_ils_deposits=None, total_usd_deposits=None):
        """
        Write a single row of account data to the Excel file without currency symbols in values.
        Apply currency formatting using Excel styles. Only the first row (Sum of Accounts) should display 'Total ILS Deposits'.
        """
        # Ensure df is converted to a DataFrame if it is a dictionary
        if isinstance(df, dict):
            df = pd.DataFrame(df).T  # Transpose to align the dictionary properly as a DataFrame

        # Get the next available row in the sheet
        next_row = sheet.max_row + 1

        # Collect data for the row (Date, Exchange Rate, and Row Type (Sum of Accounts or Specific Account))
        row_data = [date, exchange_rate, row_type]

        # Flatten the USD and ILS values from the DataFrame and add to row_data
        for tag in df.index:
            row_data.append(df.loc[tag, "USD"])  # Add USD value first
            row_data.append(df.loc[tag, "ILS"])  # Then add ILS value

        # Perform Unrealized PnL calculations for each account
        total_usd_base_value = df.loc["NetLiquidationByCurrency", "USD"] - df.loc["UnrealizedPnL", "USD"]
        unrealized_usd_pnl_percent = df.loc["UnrealizedPnL", "USD"] / total_usd_base_value

        # Add Unrealized PnL and ILS PnL fields to the row data

        # Add Unrealized USD PnL %
        row_data.append(unrealized_usd_pnl_percent)

        for total_deposits in [total_ils_deposits, total_usd_deposits]:
            if total_deposits is not None:
                # only the sum row will display ILS/USD PnL
                currency = "ILS" if total_deposits is total_ils_deposits else "USD"
                unrealized_pnl_from_deposits = df.loc["NetLiquidationByCurrency", currency] - total_deposits
                unrealized_pnl_from_deposits_percent = unrealized_pnl_from_deposits / total_deposits
                unrealized_pnl_from_deposits_in_other_currency = unrealized_pnl_from_deposits * \
                                                                 ((1 / exchange_rate) if currency == "ILS" else exchange_rate)
            else:
                unrealized_pnl_from_deposits = ""
                unrealized_pnl_from_deposits_percent = ""
                unrealized_pnl_from_deposits_in_other_currency = ""

            # Add the Total ILS Deposits only for the sum row
            if total_deposits is not None:
                row_data.append(total_deposits)
            else:
                row_data.append("")  # Leave it blank for individual accounts
    
            # Add Unrealized ILS/USD PnL fields
            row_data.append(unrealized_pnl_from_deposits)
            row_data.append(unrealized_pnl_from_deposits_percent)
            row_data.append(unrealized_pnl_from_deposits_in_other_currency)

        # Write the row data into the Excel sheet
        for col_num, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=next_row, column=col_num, value=value)
            if col_num <= 3:
                continue  # skip Date, Exchange Rate and Type cols

            subcol_type = sheet.cell(2, col_num).value
            subcol_type = subcol_type or "NIS"  # assume col is NIS by default

            if "%" in subcol_type:
                cell.style = percentage_style  # Apply percentage style to "Unrealized PnL %" columns
            elif "USD" in subcol_type:
                cell.style = usd_style  # Apply USD style for USD columns
            elif "NIS" in subcol_type:
                cell.style = nis_style  # Apply NIS style for ILS columns

            # Apply red or green font color for Unrealized PnL columns (USD and ILS)
            col_header = sheet.cell(1, col_num).value
            if not col_header:
                col_header = sheet.cell(1, col_num - 1).value  # we are at NIS subcol of PnL
            if not col_header:
                col_header = sheet.cell(1, col_num - 2).value  # we are at % subcol of PnL
            if col_header and ("Unrealized USD PnL" in col_header or "Unrealized ILS PnL" in col_header):
                if isinstance(value, (float, int)) and value != "":
                    if value < 0:
                        cell.font = Font(color="FF0000")  # Red color for negative values
                    else:
                        cell.font = Font(color="00B050")  # Green color for positive values

    @staticmethod
    def write_headers(sheet, headers):
        # Write the headers with merged cells for sub-columns
        col = 1
        for header, span in headers:
            if span > 1:
                sheet.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + span - 1)
                sheet.cell(row=1, column=col).value = header
                sheet.cell(row=1, column=col).alignment = openpyxl.styles.Alignment(horizontal='center')
                for i, subcol_name in zip(range(span), ['USD', 'NIS', '%']):
                    sheet.cell(row=2, column=col + i).value = subcol_name
            else:
                sheet.cell(row=1, column=col).value = header
                sheet.cell(row=1, column=col).alignment = openpyxl.styles.Alignment(horizontal='center')
                sheet.cell(row=2, column=col).value = ''
            col += span

    @staticmethod
    def define_excel_styles(workbook):
        # Define currency and percent styles if they don't already exist
        styles = {"usd_currency_style": "$#,##0", "nis_currency_style": "₪#,##0", "percentage_style": "0.00%"}

        for style, style_format in styles.items():
            if style not in workbook.named_styles:
                named_style = openpyxl.styles.NamedStyle(name=style, number_format=style_format)
                workbook.add_named_style(named_style)

        return ["usd_currency_style", "nis_currency_style", "percentage_style"]


# Example usage
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Process some arguments.')
    parser.add_argument('--write_to_excel', action='store_true',
                        help='Whether to write the account info to an Excel file')
    args = parser.parse_args()

    with IBAccountInfo() as account:
        sum_info, account_info = account.get_account_info(write_to_excel=args.write_to_excel)
        print("Sum of all accounts:\n", sum_info)
        print("Individual account information:\n", account_info)
