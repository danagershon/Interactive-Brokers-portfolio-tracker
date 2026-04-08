import pandas as pd
import openpyxl
from openpyxl.styles import Font
from datetime import datetime
import logging
import pathlib
from typing import Optional, Any, Callable

from utils import IbApiConstants


class AccountInfoExcelColumns:

    class MainHeader:
        """
        The main headers of the account info Excel file (first row).
        """
        DATE = "Date"
        EXCHANGE_RATE = "Exchange Rate"
        TYPE = "Type"
        NET_LIQUIDATION = "Net Liquidation"
        STOCK_MARKET_VALUE = "Stock Market Value"
        TOTAL_CASH_BALANCE = "Total Cash Balance"
        NET_DIVIDEND = "Net Dividend"
        IB_UNREALIZED_USD_PNL = "IB Unrealized USD PnL"
        TOTAL_ILS_DEPOSITS = "Total ILS Deposits"
        UNREALIZED_ILS_PNL = "Unrealized ILS PnL"

    class SubHeader:
        """
        Optional subheaders for a main header (second row). 
        Relevant for monetary values main headers like Net Liquidation, Stock Market Value, Total Cash Balance, etc.
        """
        USD = IbApiConstants.Currency.USD
        ILS = IbApiConstants.Currency.ILS
        PCT = "%"

    class StyleName:
        USD_CURRENCY_STYLE = "usd_currency_style"
        ILS_CURRENCY_STYLE = "ils_currency_style"
        PERCENTAGE_STYLE = "percentage_style"

    STYLES = {
        StyleName.USD_CURRENCY_STYLE: "$#,##0",
        StyleName.ILS_CURRENCY_STYLE: "₪#,##0",
        StyleName.PERCENTAGE_STYLE: "0.00%"
    }

    SUBHEADER_TO_STYLE_NAME = {
        SubHeader.USD: StyleName.USD_CURRENCY_STYLE,
        SubHeader.ILS: StyleName.ILS_CURRENCY_STYLE,
        SubHeader.PCT: StyleName.PERCENTAGE_STYLE,
    }

    # Define the main header row (first row)
    MAIN_HEADER_ROW = [
        MainHeader.DATE,
        MainHeader.EXCHANGE_RATE,
        MainHeader.TYPE,
        MainHeader.NET_LIQUIDATION,
        MainHeader.STOCK_MARKET_VALUE,
        MainHeader.TOTAL_CASH_BALANCE,
        MainHeader.NET_DIVIDEND,
        MainHeader.IB_UNREALIZED_USD_PNL,
        MainHeader.TOTAL_ILS_DEPOSITS,
        MainHeader.UNREALIZED_ILS_PNL,
    ]

    TOTAL_VALUE_MAIN_HEADERS = [
        MainHeader.NET_LIQUIDATION,
        MainHeader.STOCK_MARKET_VALUE,
        MainHeader.TOTAL_CASH_BALANCE,
        MainHeader.NET_DIVIDEND,
    ]

    TOTAL_VALUE_SUBHEADERS = [
        SubHeader.USD,
        SubHeader.ILS,
    ]

    TOTAL_VALUE_MAIN_HEADER_TO_ACCOUNT_BALANCE_FIELD = {
        MainHeader.NET_LIQUIDATION: IbApiConstants.AccountBalanceField.NET_LIQUIDATION_BY_CURRENCY,
        MainHeader.STOCK_MARKET_VALUE: IbApiConstants.AccountBalanceField.STOCK_MARKET_VALUE,
        MainHeader.TOTAL_CASH_BALANCE: IbApiConstants.AccountBalanceField.TOTAL_CASH_BALANCE,
        MainHeader.NET_DIVIDEND: IbApiConstants.AccountBalanceField.NET_DIVIDEND,
    }

    UNREALIZED_PNL_MAIN_HEADERS = [
        MainHeader.IB_UNREALIZED_USD_PNL,
        MainHeader.UNREALIZED_ILS_PNL,
    ]

    UNREALIZED_PNL_MAIN_HEADER_TO_SOURCE_CURRENCY = {
        MainHeader.IB_UNREALIZED_USD_PNL: IbApiConstants.Currency.USD,
        MainHeader.UNREALIZED_ILS_PNL: IbApiConstants.Currency.ILS,
    }

    UNREALIZED_PNL_SUBHEADERS = [
        SubHeader.USD,
        SubHeader.ILS,
        SubHeader.PCT,
    ]

    # Define the mapping of main headers to their subheaders (second row)
    MAIN_HEADER_TO_SUBHEADERS = {
        MainHeader.DATE: None,
        MainHeader.EXCHANGE_RATE: None,
        MainHeader.TYPE: None,
        MainHeader.NET_LIQUIDATION: TOTAL_VALUE_SUBHEADERS,
        MainHeader.STOCK_MARKET_VALUE: TOTAL_VALUE_SUBHEADERS,
        MainHeader.TOTAL_CASH_BALANCE: TOTAL_VALUE_SUBHEADERS,
        MainHeader.NET_DIVIDEND: TOTAL_VALUE_SUBHEADERS,
        MainHeader.IB_UNREALIZED_USD_PNL: UNREALIZED_PNL_SUBHEADERS,
        MainHeader.TOTAL_ILS_DEPOSITS: [SubHeader.ILS],
        MainHeader.UNREALIZED_ILS_PNL: UNREALIZED_PNL_SUBHEADERS,
    }

    @staticmethod
    def get_flat_columns() -> list[tuple[str, str]]:
        """
        Flatten the schema to the physical Excel column order.

        Returns:
            list[(main_header, subheader)] where subheader is "" for single-column headers.
        """
        cols: list[tuple[str, str]] = []
        for main_header in AccountInfoExcelColumns.MAIN_HEADER_ROW:
            sub_headers = AccountInfoExcelColumns.MAIN_HEADER_TO_SUBHEADERS[main_header]
            if not sub_headers:
                cols.append((main_header, ""))
            else:
                for sub_header in sub_headers:
                    cols.append((main_header, sub_header))
        return cols

    @staticmethod
    def should_color_pnl(main_header: "AccountInfoExcelColumns.MainHeader", value: Any) -> bool:
        return main_header in AccountInfoExcelColumns.UNREALIZED_PNL_MAIN_HEADERS and isinstance(value, (float, int))

    # Signature for per-header value getters.
    # The getter returns the flattened values matching the header's subheaders.
    ValueGetter = Callable[[dict[str, Any]], list[Any]]

    @staticmethod
    def _get_total_value_row(main_header: str, df: pd.DataFrame) -> list[Any]:
        account_balance_field = AccountInfoExcelColumns.TOTAL_VALUE_MAIN_HEADER_TO_ACCOUNT_BALANCE_FIELD[main_header]
        return [df.at[account_balance_field, currency] for currency in AccountInfoExcelColumns.TOTAL_VALUE_SUBHEADERS]

    @staticmethod
    def _get_ib_unrealized_pnl_row(df: pd.DataFrame) -> list[Any]:
        """
        IB-reported unrealized PnL in USD
        """
        net_liq_usd = df.at[IbApiConstants.AccountBalanceField.NET_LIQUIDATION_BY_CURRENCY, IbApiConstants.Currency.USD]
        pnl_usd = df.at[IbApiConstants.AccountBalanceField.UNREALIZED_PNL, IbApiConstants.Currency.USD]
        pnl_ils = df.at[IbApiConstants.AccountBalanceField.UNREALIZED_PNL, IbApiConstants.Currency.ILS]
        base_investment = net_liq_usd - pnl_usd  # cost basis approximation
        pct = pnl_usd / base_investment
        return [pnl_usd, pnl_ils, pct]

    @staticmethod
    def _get_unrealized_pnl_from_deposits_row(df: pd.DataFrame, exchange_rate: float, total_ils_deposits: float) -> list[Any]:
        """
        Unrealized ILS PnL derived from deposits.
        """
        if total_ils_deposits is None:
            return [""] * len(AccountInfoExcelColumns.UNREALIZED_PNL_SUBHEADERS)

        net_liq_ils = df.at[IbApiConstants.AccountBalanceField.NET_LIQUIDATION_BY_CURRENCY, IbApiConstants.Currency.ILS]
        pnl_ils = net_liq_ils - total_ils_deposits
        pnl_usd = round(pnl_ils * (1 / exchange_rate), 2)
        pct = pnl_ils / total_ils_deposits
        return [pnl_usd, pnl_ils, pct]

    @staticmethod
    def _get_date(inputs: dict[str, Any]) -> list[Any]:
        return [inputs.get("date", "")]

    @staticmethod
    def _get_exchange_rate(inputs: dict[str, Any]) -> list[Any]:
        return [inputs.get("exchange_rate", "")]

    @staticmethod
    def _get_row_type(inputs: dict[str, Any]) -> list[Any]:
        return [inputs.get("row_type")]

    @staticmethod
    def _get_total_ils_deposits(inputs: dict[str, Any]) -> list[Any]:
        return [inputs.get("total_ils_deposits", "")]

    @staticmethod
    def _get_total_value_for_header(main_header: str) -> "AccountInfoExcelColumns.ValueGetter":
        return lambda inputs: AccountInfoExcelColumns._get_total_value_row(main_header=main_header, df=inputs["df"])

    @staticmethod
    def _get_ib_unrealized_pnl(inputs: dict[str, Any]) -> list[Any]:
        return AccountInfoExcelColumns._get_ib_unrealized_pnl_row(df=inputs["df"])

    @staticmethod
    def _get_unrealized_pnl_from_deposits(inputs: dict[str, Any]) -> list[Any]:
        return AccountInfoExcelColumns._get_unrealized_pnl_from_deposits_row(
            df=inputs["df"], 
            exchange_rate=inputs["exchange_rate"], 
            total_ils_deposits=inputs["total_ils_deposits"]
        )

    MAIN_HEADER_TO_VALUE_GETTER: dict[str, ValueGetter] = {
        MainHeader.DATE: _get_date.__func__,
        MainHeader.EXCHANGE_RATE: _get_exchange_rate.__func__,
        MainHeader.TYPE: _get_row_type.__func__,
        MainHeader.TOTAL_ILS_DEPOSITS: _get_total_ils_deposits.__func__,

        MainHeader.NET_LIQUIDATION: _get_total_value_for_header(MainHeader.NET_LIQUIDATION),
        MainHeader.STOCK_MARKET_VALUE: _get_total_value_for_header(MainHeader.STOCK_MARKET_VALUE),
        MainHeader.TOTAL_CASH_BALANCE: _get_total_value_for_header(MainHeader.TOTAL_CASH_BALANCE),
        MainHeader.NET_DIVIDEND: _get_total_value_for_header(MainHeader.NET_DIVIDEND),

        MainHeader.IB_UNREALIZED_USD_PNL: _get_ib_unrealized_pnl.__func__,
        MainHeader.UNREALIZED_ILS_PNL: _get_unrealized_pnl_from_deposits.__func__,
    }

    @staticmethod
    def get_row_values(main_header: "AccountInfoExcelColumns.MainHeader", inputs: dict[str, Any]) -> list[Any]:
        """
        Return the *flattened* value(s) to write for a main header, aligned to its subheaders.
        """
        if main_header not in AccountInfoExcelColumns.MAIN_HEADER_ROW:
            raise ValueError(f"Invalid main header: {main_header}")
        
        value_getter = AccountInfoExcelColumns.MAIN_HEADER_TO_VALUE_GETTER.get(main_header)
        if not value_getter:
            raise ValueError(f"Unhandled main header: {main_header}")

        return value_getter(inputs)


class ExcelHelper:

    class Colors:
        RED = "FF0000"
        GREEN = "00B050"

    SUM_OF_ACCOUNTS_ROW_TYPE = "Sum of Accounts"

    def __init__(self):
        self.workbook = None
        self.sheet = None

    def write_account_info_to_excel(self, account_info_output_file: pathlib.Path, account_desc: dict[str, str], sum_df: pd.DataFrame, 
                                    account_info: dict[str, pd.DataFrame], exchange_rate: float, total_ils_deposits: float):
        """
        Write account info into an Excel file. 
        Rows will be written for: sum of accounts and individual account details.

        Args:
            account_info_output_file: pathlib.Path - the path to the Excel file to write the account info to
            account_desc: dict[str, str] - the description of the account (mapping account ID to description)
            sum_df: pd.DataFrame - the DataFrame for the sum of accounts
            account_info: dict[str, pd.DataFrame] - the DataFrame for the individual accounts
            exchange_rate: float - the current USD to ILS exchange rate
            total_ils_deposits: float - the total ILS deposits of the master account
        """
        self._load_or_create_workbook(account_info_output_file)
        self._define_excel_styles()
        curr_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Write the sum of all accounts (first row)
        self._write_row_to_excel(
            row_type=self.SUM_OF_ACCOUNTS_ROW_TYPE, df=sum_df,
            # only the sum row will have a date, exchange rate, and total ILS deposits
            date=curr_datetime, exchange_rate=exchange_rate, total_ils_deposits=total_ils_deposits
        )

        # Write individual account data (second row and onwards)
        for account_id, account_data_df in account_info.items():
            account_name = account_id + " " + account_desc[account_id]
            self._write_row_to_excel(row_type=account_name, df=account_data_df)

        # Save the workbook
        self.workbook.save(account_info_output_file)

    def _load_or_create_workbook(self, account_info_output_file: pathlib.Path):
        """Load the Excel workbook or create a new one if it doesn't exist"""
        try:
            logging.debug(f"Loading workbook from: {account_info_output_file}")
            self.workbook = openpyxl.load_workbook(account_info_output_file)
            self.sheet = self.workbook.active

            # Check if there is already data in the file (beyond headers)
            if self.sheet.max_row > 2:
                # Explicitly append a blank row by writing an empty string to each column
                self.sheet.append(["" for _ in range(self.sheet.max_column)])

        except FileNotFoundError:
            self.workbook = openpyxl.Workbook()
            self.sheet = self.workbook.active
            self._write_headers()

    def _write_row_to_excel(self, row_type: str, df: pd.DataFrame, date: str=None, exchange_rate: float=None, 
                            total_ils_deposits: float=None):
        """
        Write a single row of account data to the Excel file.

        Args:
            row_type: str - the type of the row (Sum of Accounts or Specific Account)
            df: pd.DataFrame - the DataFrame for the row (contains the account balance data)
            date: str - the current datetime
            exchange_rate: float - the USD to ILS exchange rate
            total_ils_deposits: float - the total ILS deposits of the master account
        """
        # Get the next available row in the sheet
        next_row = self.sheet.max_row + 1

        row_data = []
        inputs = {
            "df": df,
            "row_type": row_type,
            "date": date,
            "exchange_rate": exchange_rate,
            "total_ils_deposits": total_ils_deposits,
        }
        for main_header in AccountInfoExcelColumns.MAIN_HEADER_ROW:
            row_value = AccountInfoExcelColumns.get_row_values(main_header=main_header, inputs=inputs)
            row_data.extend(row_value)

        # Write the row data into the Excel sheet
        flat_cols = AccountInfoExcelColumns.get_flat_columns()
        for col_num, (value, (main_header, subheader)) in enumerate(zip(row_data, flat_cols), start=1):
            cell = self.sheet.cell(row=next_row, column=col_num, value=value)
            self._apply_cell_style(cell=cell, value=value, main_header=main_header, subheader=subheader)

    def _write_headers(self):
        """
        Write a two-row header, with merged cells for repeated groups.
        """
        group_start_col = 1
        curr_group: Optional[str] = None
        flat_cols = AccountInfoExcelColumns.get_flat_columns()

        def flush_group(end_col: int):
            nonlocal group_start_col, curr_group
            if curr_group is None:
                return
            if end_col > group_start_col:
                self.sheet.merge_cells(
                    start_row=1,
                    start_column=group_start_col,
                    end_row=1,
                    end_column=end_col,
                )
            self.sheet.cell(row=1, column=group_start_col).value = curr_group
            self.sheet.cell(row=1, column=group_start_col).alignment = openpyxl.styles.Alignment(horizontal="center")
            curr_group = None

        for idx, (header_group, subheader) in enumerate(flat_cols, start=1):
            if header_group != curr_group:
                flush_group(idx - 1)
                curr_group = header_group
                group_start_col = idx
            self.sheet.cell(row=2, column=idx).value = subheader

        flush_group(len(flat_cols))

    def _define_excel_styles(self):
        """
        Define currency and percent styles if they don't already exist
        """
        for style, style_format in AccountInfoExcelColumns.STYLES.items():
            if style not in self.workbook.named_styles:
                named_style = openpyxl.styles.NamedStyle(name=style, number_format=style_format)
                self.workbook.add_named_style(named_style)

    def _apply_cell_style(self, cell, value: Any, main_header: str, subheader: str):
        style_name = AccountInfoExcelColumns.SUBHEADER_TO_STYLE_NAME.get(subheader)
        if style_name:
            cell.style = style_name

        if AccountInfoExcelColumns.should_color_pnl(main_header, value):
            cell.font = Font(color=self.Colors.RED) if value < 0 else Font(color=self.Colors.GREEN)
